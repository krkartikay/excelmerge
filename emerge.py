import openpyxl
from fuzzywuzzy import fuzz

NULL_STR = "(null)"
WARN_STR = "⚠ "

ALGOS = [fuzz.ratio,
         fuzz.WRatio,
         fuzz.partial_token_set_ratio,
         fuzz.token_set_ratio,
         fuzz.token_sort_ratio,
         fuzz.partial_ratio,
         fuzz.partial_token_set_ratio,
         fuzz.partial_token_sort_ratio]

class ExcelFile():
    def __init__(self, fname):
        self.fname = fname
        self.startRow = 0
        self.endRow = 0
        self.startCol = 0
        self.endCol = 0
        self.mergeon = 0
        
        self.loadFile()

    def __repr__(self) -> str:
        return "<file \"%s\" (%d,%d):(%d,%d) [%d]> %s" % (
            self.fname,
            self.startRow,
            self.endRow,
            self.startCol,
            self.endCol,
            self.mergeon,
            str(self.tableData),
        )

    def setStartRow(self, startRow):
        try:
            self.startRow = int(startRow)
            self.updateData()
        except ValueError:
            pass

    def setEndRow(self, endRow):
        try:
            self.endRow = int(endRow)
            self.updateData()
        except ValueError:
            pass

    def setStartCol(self, startCol):
        if startCol:
            self.startCol = openpyxl.utils.column_index_from_string(startCol)
            self.updateData()

    def setEndCol(self, endCol):
        if endCol:
            self.endCol = openpyxl.utils.column_index_from_string(endCol)
            self.updateData()

    def setMergeon(self, mergeon):
        self.mergeon = mergeon

    def loadFile(self):
        self.tableData = []
        if self.fname == "":
            return
        wb = openpyxl.load_workbook(self.fname)
        sheet = wb.active
        for row in sheet.rows:
            self.tableData.append([str(x.value) if x.value is not None else "" for x in row])
        if len(self.tableData):
            self.startRow = 1
            self.endRow = len(self.tableData)
            self.startCol = 1
            self.endCol = len(self.tableData[0])
        self.removeEmptyRowsCols()
        self.updateData()
        wb.close()

    def removeEmptyRowsCols(self):
        n = len(self.tableData)
        m = len(self.tableData[0]) if n else 0
        firstNonEmptyRow = 0
        firstNonEmptyCol = 0
        for i in range(n):
            if any(self.tableData[i][j] for j in range(m)):
                firstNonEmptyRow = i+1
                break
        for j in range(m):
            if any(self.tableData[i][j] for i in range(n)):
                firstNonEmptyCol = j+1
                break
        self.startRow = firstNonEmptyRow
        self.startCol = firstNonEmptyCol

    def updateData(self):
        self.tableData = []
        wb = openpyxl.load_workbook(self.fname)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=self.startRow,
                                   max_row=self.endRow,
                                   min_col=self.startCol,
                                   max_col=self.endCol):
            self.tableData.append([str(x.value) if x.value is not None else "" for x in row])

class MergeManager():
    def __init__(self):
        self.file1 = ExcelFile("")
        self.file2 = ExcelFile("")
        self.fuzzy = True
        self.algo = 2
        self.thresh = 70
        self.outputType = "outerjoin"
        self.updateData()

    def setFile1(self, file1):
        self.file1 = ExcelFile(file1)

    def setFile2(self, file2):
        self.file2 = ExcelFile(file2)

    def setFuzzy(self, fuzzy):
        self.fuzzy = fuzzy

    def setAlgo(self, algo):
        self.algo = algo

    def setThresh(self, thresh):
        try:
            self.thresh = thresh
        except ValueError:
            self.thresh = 80

    def setOutputType(self, outputType):
        self.outputType = outputType

    def updateData(self):
        n1 = len(self.file1.tableData)
        m1 = len(self.file1.tableData[0]) if n1 else 0
        n2 = len(self.file2.tableData)
        m2 = len(self.file2.tableData[0]) if n2 else 0
        
        self.startRow = 1
        self.startCol = 1
        self.endRow = 1
        self.endCol = 1
        
        if not (n1 and m1 and n2 and m2):
            self.tableData = []
            return
        
        matchFxn = self.fuzzymatcher if self.fuzzy else self.exactmatcher

        self.tableData = []
        n1 = len(self.file1.tableData)
        n2 = len(self.file2.tableData)
        m1 = len(self.file1.tableData[0]) if n1 else 0
        m2 = len(self.file2.tableData[0]) if n2 else 0
        j1 = self.file1.mergeon - self.file1.startCol
        j2 = self.file2.mergeon - self.file2.startCol
        if not (j1 >= 0 and j2 >= 0 and j1 < m1 and j2 < m2):
            return
        data1 = [self.file1.tableData[i1][j1] for i1 in range(n1)]
        data2 = [self.file2.tableData[i2][j2] for i2 in range(n2)]
        
        rowsNonMatchedRight = set(range(n2))

        for l in range(n1):
            matches = 0
            rows = []
            for r in range(n2):
                if data1[l].strip() == "" or data2[r].strip() == "":
                    continue
                if matchFxn(data1[l], data2[r]):
                    matches += 1
                    row = self.file1.tableData[l] + self.file2.tableData[r]
                    rows.append(row)
                    if r in rowsNonMatchedRight:
                        rowsNonMatchedRight.remove(r)
            if matches == 0:
                row = self.file1.tableData[l] + [NULL_STR] * m2
                if self.outputType == "leftjoin" or self.outputType == "outerjoin":
                    self.tableData.append(row)
            if matches == 1:
                self.tableData.append(rows[0])
            if matches > 1:
                for r in rows:
                    r[0] = WARN_STR + r[0]
                    self.tableData.append(r)
        if self.outputType == "outerjoin":
            for r in rowsNonMatchedRight:
                row = [NULL_STR] * m1 + self.file2.tableData[r]
                self.tableData.append(row)

        self.endRow = len(self.tableData)
        self.endCol = len(self.tableData[0]) if self.endRow else 1

    def fuzzymatcher(self, d1, d2):
        algo = ALGOS[self.algo]
        return (algo(d1, d2) >= self.thresh)

    def exactmatcher(self, d1, d2):
        return str(d1).strip().lower() == str(d2).strip().lower()

    def save(self, fname):
        if not fname.endswith(".xlsx"):
            fname += ".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Merge Data"
        for row in self.tableData:
            for i in range(len(row)):
                row[i] = toVal(row[i])
            ws.append(row)
        for row in ws:
            for cell in row:
                if cell.value == NULL_STR:
                    cell.font = openpyxl.styles.Font(name='Calibri',color="FF0000")
                if str(cell.value).startswith(WARN_STR):
                    cell.font = openpyxl.styles.Font(name='Calibri',color="0000FF")
                    cell.value = str(cell.value)[2:]
        wb.save(fname)

def toVal(x):
    try:
        return int(x)
    except ValueError:
        pass
    try:
        return float(x)
    except ValueError:
        pass
    x = str(x)
    return x